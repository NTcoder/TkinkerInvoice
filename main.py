from tkinter import *
from openpyxl import Workbook
from tkinter import filedialog
from tkinter import messagebox
import datetime
from openpyxl import load_workbook
from n2w import convert
#import Tkinter.messagebox

class Application(Frame):
    def say_hi(self) :
        print ("hi there, everyone!")

    def __init__(self, master=None) :
        Frame.__init__(self, master)
        #self.pack()
        self.grid()
        self.styling_elements()
        self.createWidgets()
        

    def createWidgets(self) :

        #Create import button
        self.importReport = Button(self , height='3', width='12', font='Helvetica 18 bold')
        self.importReport["text"] = "Import Report"
        self.importReport["fg"] = self.text_colour
        self.importReport["bg"] = self.backgroud_color
        self.importReport["command"] = self.fileopener
        #self.importReport.pack(side=LEFT, pady=10, padx=10)
        self.importReport.grid(row=0, column=0, padx=15, pady=15)
        
        #Create NewReport button
        self.NewReport = Button(self, height='3', width='12',font='Helvetica 18 bold')
        self.NewReport["text"] = "New Report"
        self.NewReport["fg"] = self.text_colour
        self.NewReport["bg"] = self.backgroud_color
        self.NewReport["command"] = self.EnterFileName
        #self.NewReport.pack(side=LEFT, pady=10, padx=10)       
        self.NewReport.grid(row=0, column=1, padx=15, pady=15)

        #Create QUIT button
        self.QUIT = Button(self)
        self.QUIT["text"] = "QUIT"
        self.QUIT["fg"] = 'red'
        self.QUIT["bg"] = self.backgroud_color
        self.QUIT["command"] =  self.quit
        self.QUIT.grid(row=0, column=5, padx=15, pady=15)
        #self.QUIT.pack(side=RIGHT, pady=10, padx=10)

    def EnterFileName (self):
        self.importReport.grid_forget()
        self.NewReport.grid_forget()
        #self.QUIT.grid_forget()
        self.reprotNameVar = StringVar()
        self.EnterNameLabel = Label(self)
        self.EnterNameLabel['text'] = "Please enter report name!"
        self.EnterNameLabel['fg'] = "black"
        self.EnterNameLabel.grid(row=0, column=0)
        self.textbox = Entry(root, textvariable=self.reprotNameVar)
        self.textbox.focus_set()
        self.textbox.grid(row=1, column=0)

        #Create NewReportEnterFilename button
        self.FileNameOK = Button(self)
        self.FileNameOK["text"] = "OK"
        self.FileNameOK["fg"] = "blue"
        self.FileNameOK["command"] = self.createReport
        self.FileNameOK.grid(row=2, column=0)
        #self.QUIT.grid(row=0, column=3)     



    def fileopener (self) :
        #root.withdraw() # we don't want a full GUI, so keep the root window from appearing
        self.filename = filedialog.askopenfilename() # show an "Open" dialog box and return the path to the selected file
        print(self.filename)
        if (self.filename == None or self.filename == ' ' or self.filename == ''):
            pass
        else :
            self.importReport.grid_forget()
            self.NewReport.grid_forget()
            self.wb = load_workbook(self.filename)
            self.reprotName = self.filename
            self.ws = self.wb.active
            self.createForm()

    def createReport (self) :
        if (self.reprotNameVar.get() == None or self.reprotNameVar.get() == '' or self.reprotNameVar.get() == ' ' ):
            messagebox.showerror("Error","Report name Cannot be empty, please fill some valid name")
        else :
            self.reprotName = self.reprotNameVar.get() + '.xlsx'
            self.FileNameOK.grid_forget()
            self.EnterNameLabel.grid_forget()
            self.textbox.grid_forget()
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = 'Sheet 1'
            #ws.title = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            self.ws.cell(row=1, column=1, value="Receipt Number")
            self.ws.column_dimensions['A'].width = 10
            self.ws.cell(row=1, column=2, value="Name")
            self.ws.column_dimensions['B'].width = 20
            self.ws.cell(row=1, column=3, value="Time")
            self.ws.column_dimensions['C'].width = 15
            self.ws.cell(row=1, column=4, value="Address 1")
            self.ws.column_dimensions['D'].width = 15
            self.ws.cell(row=1, column=5, value="Address 2")
            self.ws.column_dimensions['E'].width = 15
            self.ws.cell(row=1, column=6, value="City/Village")
            self.ws.column_dimensions['F'].width = 5
            self.ws.cell(row=1, column=7, value="District")
            self.ws.column_dimensions['G'].width = 5
            self.ws.cell(row=1, column=8, value="State")
            self.ws.column_dimensions['H'].width = 5
            self.ws.cell(row=1, column=9, value="Pincode")
            self.ws.column_dimensions['I'].width = 5
            self.ws.cell(row=1, column=10, value="Amount Text")
            self.ws.column_dimensions['J'].width = 20
            self.ws.cell(row=1, column=11, value="Amount Nr")
            self.ws.column_dimensions['K'].width = 10
            self.ws.cell(row=1, column=12, value="Payment Mode")
            self.ws.column_dimensions['L'].width = 5
            self.ws.cell(row=1, column=13, value="Reference No.")
            self.ws.column_dimensions['M'].width = 10
            self.ws.cell(row=1, column=14, value="Bank Name")
            self.ws.column_dimensions['N'].width = 5
            self.ws.cell(row=1, column=15, value="Branch")
            self.ws.column_dimensions['O'].width = 5
            self.ws.cell(row=1, column=16, value="Contact Number")
            self.ws.column_dimensions['P'].width = 14
            self.ws.cell(row=1, column=17, value="Email-id")
            self.ws.column_dimensions['Q'].width = 20
            #FileName = "TestSheet.xlsx"
            self.wb.save(self.reprotName)
            self.createForm()

    def styling_elements(self):
        # Variables for Styling :
        self.backgroud_color = '#6ca6cd'
        self.text_colour = '#fffaf0'
        self.font_size_small = '30'
        self.font_size_medium = '40'
        self.font_size_big = '50'


    def createForm (self):
        self.QUIT.grid(row=0, column=3) 
        # create a Form label
        self.heading = Label(root, text="Invoice Details",fg=self.text_colour, bg=self.backgroud_color)
        # create a Receipt Number label
        self.receiptNumber = Label(root, text="Receipt Number", fg=self.text_colour, bg=self.backgroud_color)
        # create a Name label
        self.Name = Label(root, text="Name*",fg=self.text_colour, bg=self.backgroud_color)
        # create a Address1 label
        self.address1 = Label(root, text="Address 1 *",fg=self.text_colour, bg=self.backgroud_color)
        # create a Address2 label
        self.address2 = Label(root, text="Address 2", fg=self.text_colour, bg=self.backgroud_color)
        # create a City/Village label
        self.city = Label(root, text="City/Village*",fg=self.text_colour, bg=self.backgroud_color)  
        # create a District label
        self.district = Label(root, text="District*",fg=self.text_colour, bg=self.backgroud_color)
        # create a State label
        self.state = Label(root, text="State*", fg=self.text_colour, bg=self.backgroud_color) 
        # create a PINCODE label
        self.pincode = Label(root, text="Pincode",fg=self.text_colour, bg=self.backgroud_color)    
        # create a Amount text label
        self.Amount = Label(root, text="Amount*", fg=self.text_colour, bg=self.backgroud_color)
        # create a AmountWord No. lable
        self.AmountWr = Label(root, text="Amount-words",fg=self.text_colour, bg=self.backgroud_color)
        # create a Contact No. label
        self.contact_no = Label(root, text="Contact No.*",fg=self.text_colour, bg=self.backgroud_color)
        # create a Email id label
        self.email_id = Label(root, text="Email id",fg=self.text_colour, bg=self.backgroud_color)
        # create a Payment mode label
        self.payment_mode = Label(root, text="Payment Mode", fg=self.text_colour, bg=self.backgroud_color)
        # create a Reference No label
        self.reference_no = Label(root, text="Reference No./ Cheque No.",fg=self.text_colour, bg=self.backgroud_color)    
        # create a Bank Name id label
        self.bank_name = Label(root, text="Bank Name*", fg=self.text_colour, bg=self.backgroud_color)
        # create a branch label
        self.branch = Label(root, text="Branch*", fg=self.text_colour, bg=self.backgroud_color)

        #set the location of different labels
        self.heading.grid(row=0, column=1)
        self.receiptNumber.grid(row=1, column=0)
        self.Name.grid(row=2, column=0)
        self.address1.grid(row=3, column=0)
        self.address2.grid(row=4, column=0)
        self.city.grid(row=5, column=0)
        self.district.grid(row=6, column=0) 
        self.state.grid(row=7, column=0) 
        self.pincode.grid(row=8, column=0) 
        self.Amount.grid(row=9, column=0)
        self.AmountWr.grid(row=10, column=0)
        self.contact_no.grid(row=11, column=0)
        self.email_id.grid(row=12, column=0)
        self.payment_mode.grid(row=13, column=0)
        self.reference_no.grid(row=14, column=0)
        self.bank_name.grid(row=15, column=0)
        self.branch.grid(row=16, column=0)

        #set different entry point for different labels
        self.name_field = Entry(root)
        self.address1_field = Entry(root)
        self.address2_field = Entry(root)
        self.city_field = Entry(root)
        self.district_field = Entry(root)
        self.state_field = Entry(root)
        self.pincode_field = Entry(root)
        self.Amount_field = Entry(root)
        self.AmountWr_field = Entry(root)
        self.contact_no_field = Entry(root)
        self.email_id_field = Entry(root)
        #self.payment_mode_field = Entry(root)
        self.reference_no_field = Entry(root)
        self.bank_name_field = Entry(root)
        self.branch_field = Entry(root)

        #set location of different grid fields
        self.name_field.grid(row=2, column=1, ipadx="100")
        self.name_field.bind("<Return>", self.focus_set_address1_field )
        self.address1_field.grid(row=3, column=1, ipadx="100")
        self.address1_field.bind("<Return>", self.focus_set_address2_field )
        self.address2_field.grid(row=4, column=1, ipadx="100")
        self.address2_field.bind("<Return>", self.focus_set_city_field )
        self.city_field.grid(row=5, column=1, ipadx="100")
        self.city_field.bind("<Return>", self.focus_set_district_field )
        self.district_field.grid(row=6, column=1, ipadx="100")
        self.district_field.bind("<Return>", self.focus_set_state_field )
        self.state_field.grid(row=7, column=1, ipadx="100")
        self.state_field.bind("<Return>", self.focus_set_pincode_field )
        self.pincode_field.grid(row=8, column=1, ipadx="100")
        self.pincode_field.bind("<Return>", self.focus_set_Amount_field )
        self.Amount_field.grid(row=9, column=1, ipadx="100")
        self.Amount_field.bind("<Return>", self.focus_set_contact_no_field )
        self.AmountWr_field.grid(row=10, column=1, ipadx="100")
        self.contact_no_field.grid(row=11, column=1, ipadx="100")
        self.contact_no_field.bind("<Return>", self.focus_set_email_id_field )
        self.email_id_field.grid(row=12, column=1, ipadx="100")
        self.email_id_field.bind("<Return>", self.focus_set_reference_no_field )
        self.payment_mode_choices = { 'Bank Transfer (IMPS)' , 'RTGS' , 'Cheque' , 'Bank Deposit' , 'NEFT' }
        self.payment_mode_field = StringVar(root)
        self.payment_mode_dropdown = OptionMenu(root, self.payment_mode_field, *self.payment_mode_choices)
        self.payment_mode_dropdown.grid(row=13, column=1, ipadx="100")
        self.payment_mode_field.set('NEFT')
        #self.payment_mode_field.bind("<Return>", self.focus_set_reference_no_field )
        self.reference_no_field.grid(row=14, column=1, ipadx="100")
        self.reference_no_field.bind("<Return>", self.focus_set_bank_name_field )
        self.bank_name_field.grid(row=15, column=1, ipadx="100")
        self.bank_name_field.bind("<Return>", self.focus_set_branch_field )
        self.branch_field.grid(row=16, column=1, ipadx="100")

        # create a Submit Button and place into the root window
        self.submit = Button(root, text="Submit", fg="Black", bg="Grey", command=self.validate, padx=30,pady=10)
        self.submit.grid(row=17, column=1)
        # create a Clear Button and place into the root window
        self.clear_button = Button(root, text="Clear", fg="White", bg="Red", command=self.clear, padx=30,pady=10)
        self.clear_button.grid(row=17, column=4)

    def focus_set_name_field (self):
        self.name_field.focus_set()

    def focus_set_address1_field (self,event):
        self.address1_field.focus_set()

    def focus_set_address2_field (self,event):
        self.address2_field.focus_set()

    def focus_set_city_field (self,event):
        self.city_field.focus_set()

    def focus_set_district_field (self,event):
        self.district_field.focus_set()

    def focus_set_state_field (self,event):
        self.state_field.focus_set()

    def focus_set_pincode_field (self,event):
        self.pincode_field.focus_set()

    def focus_set_Amount_field  (self,event):
        self.Amount_field.focus_set()  
        
    def focus_set_contact_no_field  (self,event):
        self.AmountWr_field.set(convert(Amount_field.get()))
        self.contact_no_field.focus_set()
    
    def focus_set_email_id_field  (self,event):
        self.email_id_field.focus_set()

    def focus_set_payment_mode_field  (self,event):
        self.payment_mode_field.focus_set()

    def focus_set_reference_no_field  (self,event):
        self.reference_no_field.focus_set()
    
    def focus_set_bank_name_field  (self,event):
        self.bank_name_field.focus_set()

    def focus_set_branch_field (self,event):
        self.branch_field.focus_set()

    def validate(self):
        if ( self.name_field.get()!= '' and self.address1_field.get() != '' and self.city_field.get() != '' and self.district_field.get() != '' and self.state_field.get() != '' and self.contact_no_field.get() != '' and self.Amount_field.get() != '' and self.reference_no_field.get() != '') :
            self.insert()
        else:
            messagebox.showerror("Error", " Please check if all the mandatory fields are filled !.")

    # Function to take data from GUI 
    # window and write to an excel file
    def insert (self):
        
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        self.current_row = self.ws.max_row
        self.current_column = self.ws.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        '''
        self.ws.cell(row=1, column=1, value="Receipt Number")
        self.ws.cell(row=1, column=2, value="Name")
        self.ws.cell(row=1, column=3, value="TimeStamp")
        self.ws.cell(row=1, column=4, value="Address")
        self.ws.cell(row=1, column=5, value="Amount Text")
        self.ws.cell(row=1, column=6, value="Amount Nr")
        self.ws.cell(row=1, column=7, value="Payment Mode")
        self.ws.cell(row=1, column=8, value="Reference No.")
        self.ws.cell(row=1, column=9, value="Bank Name")
        self.ws.cell(row=1, column=10, value="Branch")
        self.ws.cell(row=1, column=11, value="Contact Number")
        self.ws.cell(row=1, column=12, value="Email-id")
        '''
        self.ws.cell(row=self.current_row + 1, column=2).value = self.name_field.get()
        self.ws.cell(row=self.current_row + 1, column=3).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        self.ws.cell(row=self.current_row + 1, column=4).value = self.address1_field.get()
        self.ws.cell(row=self.current_row + 1, column=5).value = self.address2_field.get()
        self.ws.cell(row=self.current_row + 1, column=6).value = self.city_field.get()
        self.ws.cell(row=self.current_row + 1, column=7).value = self.district_field.get()
        self.ws.cell(row=self.current_row + 1, column=8).value = self.state_field.get()
        self.ws.cell(row=self.current_row + 1, column=9).value = self.pincode_field.get()
        self.ws.cell(row=self.current_row + 1, column=10).value = self.AmountWr_field.get()
        self.ws.cell(row=self.current_row + 1, column=11).value = self.Amount_field.get()
        self.ws.cell(row=self.current_row + 1, column=12).value = self.payment_mode_field.get()
        self.ws.cell(row=self.current_row + 1, column=13).value = self.reference_no_field.get()
        self.ws.cell(row=self.current_row + 1, column=14).value = self.bank_name_field.get()
        self.ws.cell(row=self.current_row + 1, column=15).value = self.branch_field.get()
        self.ws.cell(row=self.current_row + 1, column=16).value = self.contact_no_field.get()
        self.ws.cell(row=self.current_row + 1, column=17).value = self.email_id_field.get()

        # save the file
        try:
            self.wb.save(self.reprotName)
            self.clear()
        except PermissionError as e :
            messagebox.showerror("Error", """The excel file in which you are trying to save is currently open. 
            If problem persists, Please report the issue with screenshot to navadiatarun@gmail.com: """ + str(e))

    def clear(self):
        
        # clear the content of text entry box
        self.name_field.delete(0, END)
        self.address1_field.delete(0, END)
        self.address2_field.delete(0, END)
        self.city_field.delete(0, END)
        self.district_field.delete(0, END)
        self.state_field.delete(0, END)
        self.pincode_field.delete(0, END)
        self.AmountWr_field.delete(0, END)
        self.Amount_field.delete(0, END)
        self.payment_mode_field.set('NEFT')
        self.reference_no_field.delete(0, END)
        self.bank_name_field.delete(0, END)
        self.branch_field.delete(0, END)
        self.contact_no_field.delete(0, END)
        self.email_id_field.delete(0, END)
        self.focus_set_name_field()
 

root = Tk()
root.title("Swarved Mahamandir Trust Invoicing Software")
root.configure(background='#87ceeb')
#w, h = root.winfo_screenwidth(), root.winfo_screenheight()
#root.geometry("%dx%d+0+0" % (w, h))
root.geometry("800x600")
app = Application(master=root)
app.mainloop()
root.destroy()